from tkinter import *
from PIL import Image
from openpyxl import *
import pandas as pd
import random
from datetime import *

read_file = pd.read_csv (r'C:\Users\ferri\PycharmProjects\midterm\exam1.csv')
read_file.to_excel (r'C:\Users\ferri\PycharmProjects\midterm\midtermfile.xlsx', index = None, header=True)



#Window stuff
window = Tk()
window.title('CSUMB App')
window.geometry("750x800")
window.resizable(0,0)
window.configure(bg="lightblue")
img3 = PhotoImage(file="newlogo2.png")
label =Label(window, image=img3)
label.place(x=250, y=0)
now = datetime.now()



# Transparent Img Stuff
img = Image.open("logo1.jpg")
img1 = img.convert("RGBA")
img2 = img1.getdata()

transparency = []
for item in img2:
        if item[:3] == (255,255,255):
            transparency.append((255,255,255,0))
        else:
            transparency.append(item)

img1.putdata(transparency)
img1.save("newlogo2.png", "PNG")


#Excel stuff
wb = load_workbook("midtermfile.xlsx")
ws = wb.active
column1 = ws['A']
column2 = ws['B']
column3 = ws['C']

#Lists/Buttons stuff
def submit1():
    list = ''
    for cell in column1:
        list = f'{list + str(cell.value)}\n'
        label2.config(text=list)

button1= Button(window, text ="Buildings", font=("Arial", 16), fg='darkgreen', bg='lightgreen', command=submit1)
button1.place(x=175,y=200)


def submit2():
    list = ''
    for cell in column2:
        list = f'{list + str(cell.value)}\n'
        label2.config(text=list)

button1= Button(window, text ="Calendar", font=("Arial", 16), fg='darkgreen', bg='lightgreen', command=submit2)
button1.place(x=290,y=200)


def submit3():
    list = ''
    for cell in column3:
        list = f'{list + str(cell.value)}\n'
        label2.config(text=list)

button1= Button(window, text ="Faculty", font=("Arial", 16), fg='darkgreen', bg='lightgreen', command=submit3)
button1.place(x=405,y=200)

#Random stuff
def randomname(): #randomizes responses and prints them
    range = ws['C2:C10']
    name = []
    for items in range:
        for subitems in items:
            name.append(subitems.value)
        random2 = random.choice(name)
        label2.configure(text=f'{str(random2) + " wins free parking for the month of " + now.strftime("%B") + "!"}')


button1= Button(window, text ="Lucky Faculty", font=("Arial", 16), fg='darkgreen', bg='lightgreen', command=randomname)
button1.place(x=500,y=200)

label2 = Label(window, text ="", font=("Arial", 14), fg ="green", bg="white", width=40)
label2.place(x=190, y=250)

window.mainloop()